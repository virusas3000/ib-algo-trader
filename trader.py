"""
IB Algo Trader — Multi-Strategy Day Trading Engine
Strategies: ORB, VWAP Mean Reversion, Momentum, Gap Fill, Power Hour
"""
from __future__ import annotations

import argparse
import csv
import logging
import os
from pathlib import Path
import sys
import time as _time
from datetime import datetime, time, timezone
from pathlib import Path
from zoneinfo import ZoneInfo

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
from openpyxl.utils import get_column_letter
from ib_insync import IB, Stock, Index, MarketOrder, LimitOrder, util

import config as cfg
from indicators import compute_indicators
from strategy import Strategy, Signal, Side, ExitReason, StrategyName
from telegram_notifier import TelegramNotifier
from sentiment_integration import integrate_sentiment_with_trader
from ml_integration import integrate_ml_with_trader

ET = ZoneInfo("America/New_York")

# ─── Logging ─────────────────────────────────────────────
log_path = Path(__file__).parent / "algo.log"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.FileHandler(log_path),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("trader")


# ─── Trade Logger ─────────────────────────────────────────

class TradeLogger:
    HEADERS = ["Date", "Time (ET)", "Symbol", "Side", "Strategy",
               "Entry Price", "Exit Price", "Size", "P&L ($)", "P&L (%)", "Reason"]

    def __init__(self):
        self.csv_path = Path(__file__).parent / cfg.TRADE_LOG_CSV
        self.xlsx_path = Path(__file__).parent / "trade_history.xlsx"
        self._init_csv()
        self._init_xlsx()

    def _init_csv(self):
        if not self.csv_path.exists():
            with open(self.csv_path, "w", newline="") as f:
                csv.writer(f).writerow(self.HEADERS)

    def _init_xlsx(self):
        if self.xlsx_path.exists():
            return
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Trades"
        self._write_header(ws)
        # Summary sheet
        ws2 = wb.create_sheet("Summary")
        ws2["A1"] = "Daily Summary"
        ws2["A1"].font = Font(bold=True, size=14)
        wb.save(self.xlsx_path)

    def _write_header(self, ws):
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        thin = XlSide(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for col, h in enumerate(self.HEADERS, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
            cell.border = border
        # Column widths
        widths = [12, 10, 8, 6, 18, 12, 12, 8, 12, 10, 12]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def log(self, **kwargs):
        # Write to CSV
        row_data = [
            kwargs.get("datetime", "")[:10],
            kwargs.get("datetime", "")[11:19],
            kwargs.get("symbol", ""),
            kwargs.get("side", ""),
            kwargs.get("strategy", ""),
            kwargs.get("entry", ""),
            kwargs.get("exit", ""),
            kwargs.get("size", ""),
            kwargs.get("pnl", ""),
            round(kwargs.get("pnl", 0) / max(abs(kwargs.get("entry", 1) * kwargs.get("size", 1)), 1) * 100, 2),
            kwargs.get("reason", ""),
        ]
        with open(self.csv_path, "a", newline="") as f:
            csv.writer(f).writerow(row_data)

        # Write to Excel
        try:
            wb = openpyxl.load_workbook(self.xlsx_path)
            ws = wb["Trades"]
            next_row = ws.max_row + 1
            pnl = kwargs.get("pnl", 0)
            green = PatternFill("solid", fgColor="C6EFCE")
            red   = PatternFill("solid", fgColor="FFC7CE")
            thin  = XlSide(style="thin", color="CCCCCC")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(row=next_row, column=col, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal="center")
                if col == 9:  # P&L column
                    cell.fill = green if pnl >= 0 else red
                    cell.font = Font(bold=True, color="006100" if pnl >= 0 else "9C0006")
            # Update summary sheet
            self._update_summary(wb, ws)
            wb.save(self.xlsx_path)
        except Exception as e:
            log.warning(f"Excel write error: {e}")

    def _update_summary(self, wb, ws):
        try:
            ws2 = wb["Summary"]
            ws2.delete_rows(1, ws2.max_row)
            ws2["A1"] = "Daily Trading Summary"
            ws2["A1"].font = Font(bold=True, size=14, color="1F4E79")
            # Collect stats
            trades, wins, losses, total_pnl = 0, 0, 0, 0.0
            strategy_stats = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[0]:
                    continue
                trades += 1
                pnl = row[8] or 0
                total_pnl += pnl
                if pnl >= 0: wins += 1
                else: losses += 1
                strat = row[4] or "UNKNOWN"
                if strat not in strategy_stats:
                    strategy_stats[strat] = {"trades": 0, "pnl": 0.0, "wins": 0}
                strategy_stats[strat]["trades"] += 1
                strategy_stats[strat]["pnl"] += pnl
                if pnl >= 0: strategy_stats[strat]["wins"] += 1
            # Write summary
            ws2["A3"] = "Total Trades"; ws2["B3"] = trades
            ws2["A4"] = "Wins";         ws2["B4"] = wins
            ws2["A5"] = "Losses";       ws2["B5"] = losses
            ws2["A6"] = "Win Rate";     ws2["B6"] = f"{wins/trades*100:.1f}%" if trades else "0%"
            ws2["A7"] = "Total P&L";    ws2["B7"] = round(total_pnl, 2)
            ws2["B7"].font = Font(bold=True, color="006100" if total_pnl >= 0 else "9C0006")
            r = 9
            ws2.cell(r, 1, "Strategy").font = Font(bold=True)
            ws2.cell(r, 2, "Trades").font = Font(bold=True)
            ws2.cell(r, 3, "Win Rate").font = Font(bold=True)
            ws2.cell(r, 4, "P&L").font = Font(bold=True)
            for strat, s in strategy_stats.items():
                r += 1
                ws2.cell(r, 1, strat)
                ws2.cell(r, 2, s["trades"])
                ws2.cell(r, 3, f"{s['wins']/s['trades']*100:.0f}%" if s["trades"] else "0%")
                ws2.cell(r, 4, round(s["pnl"], 2))
        except Exception as e:
            log.warning(f"Summary update error: {e}")


# ─── Data Manager ─────────────────────────────────────────

class DataManager:
    def __init__(self, ib: IB):
        self.ib = ib
        self._contracts = {}

    def qualify(self, symbol: str) -> bool:
        if symbol in self._contracts:
            return True
        contract = Stock(symbol, "SMART", "USD")
        try:
            self.ib.qualifyContracts(contract)
            self._contracts[symbol] = contract
            return True
        except Exception as e:
            log.warning(f"Could not qualify {symbol}: {e}")
            return False

    def fetch_bars(self, symbol: str, duration: str = "1 D", bar_size: str = "5 mins") -> pd.DataFrame:
        if symbol not in self._contracts:
            return pd.DataFrame()
        contract = self._contracts[symbol]
        # Try several whatToShow values — paper accts often lack TRADES data
        for what in ("TRADES", "MIDPOINT", "BID_ASK"):
            try:
                bars = self.ib.reqHistoricalData(
                    contract,
                    endDateTime="",
                    durationStr=duration,
                    barSizeSetting=bar_size,
                    whatToShow=what,
                    useRTH=True,
                    formatDate=1,
                )
                if bars:
                    df = util.df(bars)
                    df.columns = [c.lower() for c in df.columns]
                    if what != "TRADES":
                        log.debug(f"{symbol} bars via {what} fallback")
                    return compute_indicators(df)
            except Exception as e:
                log.debug(f"{symbol} {what} failed: {e}")
                continue
        return pd.DataFrame()

    def fetch_vix(self) -> float:
        try:
            vix = Index("VIX", "CBOE")
            self.ib.qualifyContracts(vix)
            ticker = self.ib.reqMktData(vix, "", False, False)
            self.ib.sleep(2)
            val = ticker.last or ticker.close or 20.0
            self.ib.cancelMktData(vix)
            return float(val)
        except Exception:
            return 20.0  # default neutral


# ─── Order Manager ────────────────────────────────────────

class OrderManager:
    def __init__(self, ib: IB, dry_run: bool = False, telegram: "TelegramNotifier" = None):
        self.ib = ib
        self.dry_run = dry_run
        self.telegram = telegram

    def _on_fill(self, trade, fill):
        """Callback fired by ib_insync when an order gets a fill."""
        try:
            symbol = trade.contract.symbol
            action = trade.order.action
            avg_price = trade.orderStatus.avgFillPrice
            filled = trade.orderStatus.filled
            order_id = trade.order.orderId
            log.info(f"[FILL] {symbol} {action} {filled} @ {avg_price:.2f} (orderId={order_id})")
            if self.telegram:
                self.telegram.send_order_filled(
                    symbol=symbol,
                    action=action,
                    size=filled,
                    avg_fill_price=avg_price,
                    order_id=order_id,
                )
        except Exception as e:
            log.error(f"Error in fill callback: {e}")

    def _check_margin_cushion(self, signal: Signal) -> tuple[bool, str]:
        """Live margin pre-check. Returns (ok, reason).

        Skips entry if Available Funds < (estimated trade margin + cushion %).
        Uses naive est: notional × INIT_MARGIN_PCT (default 50% reg-T) for stocks.
        """
        try:
            self.ib.reqAccountSummary()
            self.ib.sleep(1)
            vals = {v.tag: v.value for v in self.ib.accountSummary()}
            avail = float(vals.get("AvailableFunds", 0))
            equity = float(vals.get("EquityWithLoanValue", 0))
            init_req = float(vals.get("InitMarginReq", 0))
            if avail <= 0 or equity <= 0:
                return True, "no margin data"  # fail open
            cushion_pct = float(getattr(cfg, 'MARGIN_CUSHION_PCT', 0.10))
            init_margin_pct = float(getattr(cfg, 'INIT_MARGIN_PCT', 0.50))
            est_trade_margin = signal.size * signal.entry * init_margin_pct
            required_cushion = equity * cushion_pct
            remaining_after = avail - est_trade_margin
            if remaining_after < required_cushion:
                return False, (
                    f"avail={avail:,.0f} est_margin={est_trade_margin:,.0f} "
                    f"remaining={remaining_after:,.0f} < cushion={required_cushion:,.0f} "
                    f"({cushion_pct:.0%} of equity {equity:,.0f})"
                )
            return True, (f"avail={avail:,.0f} est_margin={est_trade_margin:,.0f} "
                          f"remaining={remaining_after:,.0f}")
        except Exception as e:
            log.warning(f"[MARGIN-CHECK] failed, skipping check: {e}")
            return True, f"check failed: {e}"

    def enter(self, signal: Signal):
        # Trade-quality gate: veto entries with low predicted win-prob
        try:
            from trade_quality_gate import predict_win_prob
            from datetime import datetime, timezone, timedelta
            ET = timezone(timedelta(hours=-4))  # EDT
            now_et = datetime.now(ET)
            min_p = float(getattr(cfg, 'TRADE_QUALITY_MIN_WIN_PROB', 0.50))
            p_win = predict_win_prob(signal.symbol, signal.side.value,
                                     signal.strategy.value, now_et)
            if p_win < min_p:
                log.info(
                    f"[QUALITY-GATE] SKIP {signal.symbol} {signal.side.value} "
                    f"{signal.strategy.value}  p_win={p_win:.2f} < {min_p:.2f}"
                )
                return
            log.info(f"[QUALITY-GATE] PASS {signal.symbol}  p_win={p_win:.2f}")
        except Exception as e:
            log.warning(f"[QUALITY-GATE] disabled (model unavailable): {e}")

        # Live margin cushion check (hard gate)
        ok, msg = self._check_margin_cushion(signal)
        if not ok:
            log.warning(f"[MARGIN-GATE] SKIP {signal.symbol} {signal.side.value} — {msg}")
            if self.telegram:
                try:
                    self.telegram.send_message(
                        f"⚠️ MARGIN BLOCK\n"
                        f"{signal.symbol} {signal.side.value} x{signal.size} @ {signal.entry:.2f}\n"
                        f"{msg}"
                    )
                except Exception:
                    pass
            return
        log.info(f"[MARGIN-GATE] PASS {signal.symbol} — {msg}")

        if self.dry_run:
            log.info(f"[DRY-RUN] Would enter: {signal.symbol} {signal.side.value} x{signal.size} @ {signal.entry:.2f}")
            return
        action = "BUY" if signal.side == Side.LONG else "SELL"
        order = MarketOrder(action, signal.size)
        order.tif = "DAY"
        order.outsideRth = False
        contract = Stock(signal.symbol, "SMART", "USD")
        trade = self.ib.placeOrder(contract, order)
        trade.fillEvent += self._on_fill
        log.info(f"Order placed: {trade}")
        
        # Send Telegram trade entry notification
        if self.telegram:
            try:
                # Get approximate account size for position % calculation
                account_size = 1000000  # Default fallback, will be updated from strategy if available
                if hasattr(self, '_account_size'):
                    account_size = self._account_size
                    
                self.telegram.send_trade_entry(
                    symbol=signal.symbol,
                    side=signal.side.value,
                    strategy=signal.strategy.value,
                    entry_price=signal.entry,
                    size=signal.size,
                    account_size=account_size
                )
            except Exception as e:
                log.error(f"Failed to send trade entry notification: {e}")

    def exit(self, symbol: str, side: Side, size: int, reason: str):
        if self.dry_run:
            log.info(f"[DRY-RUN] Would exit: {symbol} {side.value} x{size} reason={reason}")
            return
        action = "SELL" if side == Side.LONG else "BUY"
        order = MarketOrder(action, size)
        order.tif = "DAY"
        order.outsideRth = False
        contract = Stock(symbol, "SMART", "USD")
        trade = self.ib.placeOrder(contract, order)
        trade.fillEvent += self._on_fill
        log.info(f"Exit order placed: {trade}")


# ─── Main Trader ──────────────────────────────────────────

class Trader:
    def __init__(self, dry_run: bool = False):
        self.dry_run = dry_run
        self.ib = IB()
        self.data: DataManager = None
        self.orders: OrderManager = None
        self.strategy: Strategy = None
        self.trade_log = TradeLogger()
        self.telegram = TelegramNotifier(cfg.TELEGRAM_BOT_TOKEN, cfg.TELEGRAM_CHAT_ID)
        self.connected = False

    def connect(self, retry: bool = False):
        max_attempts = 9999 if retry else 1
        for attempt in range(1, max_attempts + 1):
            log.info(f"Connecting to IB at {cfg.IB_HOST}:{cfg.IB_PORT} (clientId={cfg.CLIENT_ID}, attempt={attempt})...")
            try:
                if self.ib.isConnected():
                    self.ib.disconnect()
                    _time.sleep(1)
                self.ib.connect(cfg.IB_HOST, cfg.IB_PORT, clientId=cfg.CLIENT_ID, timeout=15)
                self.connected = True
                log.info("Connected to IB TWS.")
                # Allow delayed data when live subscription unavailable (paper accts)
                # 1=live, 2=frozen, 3=delayed, 4=delayed-frozen. 3 falls back gracefully.
                try:
                    self.ib.reqMarketDataType(3)
                    log.info("Market data type set to 3 (delayed when live unavailable).")
                except Exception as e:
                    log.warning(f"reqMarketDataType failed: {e}")
                if attempt > 1:  # Only send reconnection alerts, not initial connection
                    self.telegram.send_connection_alert("CONNECTED", "Successfully reconnected to IB TWS")
                return
            except Exception as e:
                log.error(f"Connection failed: {e}")
                if attempt == max_attempts:
                    self.telegram.send_connection_alert("CONNECTION FAILED", f"Unable to connect after {max_attempts} attempts")
                    sys.exit(1)
                if attempt == 1:  # Send disconnect alert on first failure
                    self.telegram.send_connection_alert("DISCONNECTED", "Lost connection to IB TWS, attempting to reconnect...")
                log.info("Retrying in 30 seconds...")
                _time.sleep(30)

    def _init_strategy(self):
        accounts = self.ib.managedAccounts()
        log.info(f"Accounts: {accounts}")
        self.ib.reqAccountSummary()
        self.ib.sleep(3)
        account_values = {v.tag: v.value for v in self.ib.accountSummary()
                         if v.tag == "NetLiquidation"}
        account_size = float(account_values.get("NetLiquidation", 100000))
        log.info(f"Account size: ${account_size:,.2f}")

        vix = self.data.fetch_vix()
        log.info(f"VIX: {vix:.1f} → Regime: {'High vol (ORB only)' if vix > cfg.VIX_HIGH else 'Low vol (mean rev)' if vix < cfg.VIX_LOW else 'Normal (all strategies)'}")

        self.strategy = Strategy(account_size, vix)
        
        # 🚀 Initialize sentiment analysis integration
        log.info("🧠 Initializing sentiment analysis...")
        try:
            self.sentiment_integration = integrate_sentiment_with_trader(self)
        except Exception as e:
            log.error(f"Sentiment integration failed: {e}")
            self.sentiment_integration = None
            log.info("📊 Continuing without sentiment analysis...")

        # 🤖 Initialize ML integration  
        log.info("🤖 Initializing ML analysis...")
        try:
            self.ml_integration = integrate_ml_with_trader(self, self.sentiment_integration)
        except Exception as e:
            log.error(f"ML integration failed: {e}")
            self.ml_integration = None
            log.info("📈 Continuing without ML analysis...")

    def run(self):
        watchlist = list(cfg.DEFAULT_WATCHLIST)
        # === Merge in trending tickers from Reddit (auto-refreshed by cron) ===
        try:
            from pathlib import Path as _P
            tr_file = _P(__file__).parent / 'trending_watchlist.txt'
            if tr_file.exists():
                trending = []
                for line in tr_file.read_text().splitlines():
                    line = line.strip()
                    if not line or line.startswith('#'):
                        continue
                    sym = line.split()[0].split('#')[0].strip()
                    if sym and sym not in watchlist:
                        trending.append(sym)
                if trending:
                    watchlist.extend(trending)
                    log.info(f"[TRENDING] Added {len(trending)} trending tickers from Reddit: {trending}")
        except Exception as e:
            log.warning(f"[TRENDING] Could not load trending list: {e}")
        # === Apply learned blacklist from historical losing trades ===
        try:
            from pathlib import Path as _P
            bl_file = _P(__file__).parent / 'symbol_blacklist.txt'
            if bl_file.exists():
                blacklist = {s.strip() for s in bl_file.read_text().splitlines() if s.strip()}
                before = len(watchlist)
                watchlist = [s for s in watchlist if s not in blacklist]
                log.info(f"[LEARN] Blacklist applied: removed {before-len(watchlist)} symbols {sorted(blacklist)}")
        except Exception as e:
            log.warning(f"[LEARN] Could not apply blacklist: {e}")
        self.data = DataManager(self.ib)
        self.orders = OrderManager(self.ib, dry_run=self.dry_run, telegram=self.telegram)
        self._init_strategy()

        log.info(f"Watchlist: {watchlist}")
        log.info(f"Strategies active: {[s.value for s in self.strategy.active_strategies]}")
        log.info("=" * 60)

        # Send startup notification
        self.telegram.send_system_start(
            account_size=self.strategy.account_size,
            watchlist=watchlist,
            strategies=[s.value for s in self.strategy.active_strategies]
        )

        # Pre-qualify contracts
        for sym in watchlist:
            if self.data.qualify(sym):
                log.info(f"  Qualified: {sym}")

        # Fetch previous closes for gap detection
        self._fetch_prev_closes(watchlist)

        # FIX 2: Startup cleanup — close any overnight positions left from previous session
        try:
            from ib_insync import MarketOrder, Stock
            overnight = [p for p in self.ib.positions() if p.position != 0]
            # Only close US stocks — don't touch HK bot's SEHK/HKD positions
            overnight = [p for p in overnight if p.contract.currency == "USD"]
            if overnight:
                log.warning(f"[STARTUP] Found {len(overnight)} overnight position(s) — closing before trading begins")
                self.telegram.send_message(f"⚠️ [STARTUP CLEANUP] Found {len(overnight)} overnight position(s) left from previous session — closing now: {[p.contract.symbol for p in overnight]}")
                for p in overnight:
                    try:
                        sym = p.contract.symbol
                        qty = abs(int(p.position))
                        action = "SELL" if p.position > 0 else "BUY"
                        # Use the actual contract from IB (preserves exchange/currency for non-US stocks)
                        contract = p.contract
                        self.ib.qualifyContracts(contract)
                        order = MarketOrder(action, qty)
                        order.tif = "GTC"  # avoid DAY preset cancellation
                        order.outsideRth = True
                        self.ib.placeOrder(contract, order)
                        log.info(f"[STARTUP] Closed overnight: {action} {qty} {sym} ({contract.exchange}/{contract.currency})")
                    except Exception as e:
                        log.error(f"[STARTUP] Failed to close overnight {p.contract.symbol}: {e}")
                self.ib.sleep(5)
                remaining = [p for p in self.ib.positions() if p.position != 0]
                if remaining:
                    log.error(f"[STARTUP] Still {len(remaining)} positions after cleanup: {[p.contract.symbol for p in remaining]}")
                    self.telegram.send_message(f"🚨 [STARTUP] Could not close all overnight positions: {[p.contract.symbol for p in remaining]}")
                else:
                    log.info("[STARTUP] ✅ All overnight positions cleared")
                    self.telegram.send_message("✅ [STARTUP] All overnight positions cleared — ready to trade")
        except Exception as e:
            log.error(f"[STARTUP] Overnight cleanup failed: {e}")

        log.info(f"Waiting for market... ET time: {datetime.now(ET).strftime('%H:%M:%S ET')}")

        try:
            while True:
                self.ib.sleep(1)  # keeps ib_insync event loop alive
                if not self.ib.isConnected():
                    log.warning("[RECONNECT] Lost connection — reconnecting...")
                    self.connect(retry=True)
                    self.data = DataManager(self.ib)
                    self.orders = OrderManager(self.ib, dry_run=self.dry_run, telegram=self.telegram)
                    for s in watchlist:
                        self.data.qualify(s)
                    log.info("[RECONNECT] Back online!")
                    # FIX 1: After reconnect, if we're past EOD time, immediately flatten all
                    reconnect_t = datetime.now(ET).time()
                    if reconnect_t >= time(cfg.FORCE_CLOSE_HOUR, cfg.FORCE_CLOSE_MIN):
                        log.warning("[RECONNECT EOD] Reconnected after EOD — force-closing all positions NOW")
                        self.telegram.send_message("🔒 [RECONNECT EOD] Bot reconnected after market close — force-closing all open positions")
                        self._flatten_all(ExitReason.EOD_CLOSE)
                        self._print_summary()
                        self._send_daily_summary()
                        break
                now = datetime.now(ET)
                now_t = now.time()

                if now_t < time(cfg.MARKET_OPEN_HOUR, cfg.MARKET_OPEN_MIN):
                    if now.second == 0 and now.minute % 5 == 0:
                        log.info(f"Pre-market... {now.strftime('%H:%M ET')}")
                    continue

                # Day Trading Enforcement Checks
                if self.strategy.is_daily_loss_exceeded():
                    log.error("KILL SWITCH — Max daily loss exceeded! Closing all.")
                    self._flatten_all(ExitReason.MAX_LOSS)
                    break

                if now_t >= time(cfg.FINAL_WARNING_HOUR, cfg.FINAL_WARNING_MIN) and not hasattr(self, '_final_warning_sent'):
                    if self.strategy.positions:
                        log.warning("[DAY TRADE] Final warning: Force close in 5 minutes!")
                        self.telegram.send_message(f"🚨 DAY TRADE WARNING: {len(self.strategy.positions)} positions will be force-closed in 5 minutes at {cfg.FORCE_CLOSE_HOUR}:{cfg.FORCE_CLOSE_MIN:02d} PM ET")
                        self._final_warning_sent = True

                if now_t >= time(cfg.FORCE_CLOSE_HOUR, cfg.FORCE_CLOSE_MIN):
                    if self.strategy.positions:
                        log.info("[DAY TRADE EOD] Force closing all positions for day trading compliance...")
                        self.telegram.send_message(f"🔒 DAY TRADE: Closing all {len(self.strategy.positions)} positions at market close")
                        self._flatten_all(ExitReason.EOD_CLOSE)
                    else:
                        log.info("[DAY TRADE EOD] No positions to close - day trading session complete")
                    self._print_summary()
                    self._send_daily_summary()
                    break

                # Prevent new positions near market close
                can_open_new_positions = now_t < time(cfg.NO_NEW_POSITIONS_HOUR, cfg.NO_NEW_POSITIONS_MIN)
                if not can_open_new_positions and not hasattr(self, '_new_pos_warning_sent'):
                    log.info(f"[DAY TRADE] No new positions after {cfg.NO_NEW_POSITIONS_HOUR}:{cfg.NO_NEW_POSITIONS_MIN:02d} PM ET")
                    self._new_pos_warning_sent = True

                # Check for aggressive trading mode
                aggressive_mode = Path("aggressive_mode.flag").exists()
                scan_frequency = 1 if aggressive_mode else 2  # Every minute if aggressive, every 2 minutes otherwise
                
                if now.second != 0:  # Only scan at top of minute
                    continue

                # More frequent scanning in aggressive mode
                should_scan = now.minute % scan_frequency == 0
                
                if should_scan:
                    log.info(f"[SCAN] {now.strftime('%H:%M ET')} | "
                             f"Positions: {len(self.strategy.positions)} | "
                             f"Trades: {self.strategy.trade_count} | "
                             f"P&L: ${self.strategy.daily_pnl:+.2f}")
                    
                    # 🧠 Check sentiment opportunities every 10 minutes
                    if hasattr(self, 'sentiment_integration') and now.minute % 10 == 0:
                        try:
                            self.check_sentiment_opportunities()
                            if now.minute % 30 == 0:  # Log sentiment summary every 30 min
                                self.sentiment_integration.log_sentiment_summary()
                        except Exception as e:
                            log.error(f"Sentiment opportunity check failed: {e}")

                    # 🤖 ML analysis and retraining every 15 minutes
                    if hasattr(self, 'ml_integration') and now.minute % 15 == 0:
                        try:
                            self.check_ml_opportunities()
                            if now.minute % 60 == 0:  # Log ML status every hour
                                self.ml_integration.log_ml_summary()
                                self.manage_ml_retraining()  # Check for retraining
                        except Exception as e:
                            log.error(f"ML analysis failed: {e}")

                for sym in watchlist:
                    try:
                        self._process_symbol(sym, now)
                    except Exception as e:
                        log.error(f"Error on {sym}: {e}")
                        self.telegram.send_error_alert(str(e), sym)

        except KeyboardInterrupt:
            log.info("Interrupted.")
        finally:
            self._cleanup()

    def _process_symbol(self, symbol: str, now: datetime):
        df = self.data.fetch_bars(symbol)
        if df.empty:
            return

        # Update SPY open price for momentum bias
        if symbol == "SPY" and self.strategy.spy_open is None:
            self.strategy.spy_open = float(df.iloc[0]["open"])

        # Update ORB levels
        for bar in df.itertuples():
            self.strategy.update_orb(
                symbol,
                pd.Timestamp(bar.date).to_pydatetime().replace(tzinfo=ET),
                float(bar.high), float(bar.low)
            )

        # Check exits first
        exit_reason = self.strategy.check_exits(symbol, df)
        if exit_reason:
            pos = self.strategy.positions[symbol]
            exit_price = float(df.iloc[-1]["close"])
            self.orders.exit(symbol, pos.side, pos.size, exit_reason.value)
            self.strategy.record_exit(symbol, exit_price, exit_reason)
            
            # Log trade and send Telegram notification
            pnl = round((exit_price - pos.entry_price) * pos.size * (1 if pos.side == Side.LONG else -1), 2)
            self.trade_log.log(
                datetime=now.isoformat(), symbol=symbol,
                side=pos.side.value, strategy=pos.strategy.value,
                entry=pos.entry_price, exit=exit_price,
                size=pos.size, pnl=pnl,
                reason=exit_reason.value
            )
            
            # Send Telegram exit notification
            self.telegram.send_trade_exit(
                symbol=symbol, side=pos.side.value, strategy=pos.strategy.value,
                entry_price=pos.entry_price, exit_price=exit_price, 
                size=pos.size, pnl=pnl, reason=exit_reason.value
            )
            return

        # Check entries — try all strategies (only if day trading allows new positions)
        now_t = now.time()
        can_open_new_positions = now_t < time(cfg.NO_NEW_POSITIONS_HOUR, cfg.NO_NEW_POSITIONS_MIN)
        
        if not can_open_new_positions:
            return  # Skip entry checks after cutoff time for day trading
            
        signal = None
        for check_fn in [
            lambda: self.strategy.check_gap_fill(symbol, df, now),
            lambda: self.strategy.check_orb(symbol, df, now),
            lambda: self.strategy.check_vwap_reversion(symbol, df, now),
            lambda: self.strategy.check_momentum(symbol, df, now),
            lambda: self.strategy.check_power_hour(symbol, df, now),
            lambda: self.strategy.check_vwap_reclaim(symbol, df, now),
            lambda: self.strategy.check_gap_and_go(symbol, df, now),
            lambda: self.strategy.check_bull_flag(symbol, df, now),
            lambda: self.strategy.check_rsi_extreme(symbol, df, now),
            lambda: self.strategy.check_hod_breakout(symbol, df, now),
            lambda: self.strategy.check_ema_cross(symbol, df, now),
            # Andrew Aziz strategies
            lambda: self.strategy.check_abcd_pattern(symbol, df, now),
            lambda: self.strategy.check_red_to_green(symbol, df, now),
            lambda: self.strategy.check_bottom_reversal(symbol, df, now),
            lambda: self.strategy.check_fallen_angel(symbol, df, now),
            lambda: self.strategy.check_ma_trend(symbol, df, now),
        ]:
            signal = check_fn()
            if signal:
                break

        if signal:
            # Cooldown check before placing order
            last_exit = self.strategy._exit_cooldown.get(signal.symbol)
            if last_exit:
                elapsed = (datetime.now(ET) - last_exit.replace(tzinfo=ET) if last_exit.tzinfo is None else datetime.now(ET) - last_exit).total_seconds() / 60
                if elapsed < 30:
                    log.info(f"[COOLDOWN] Blocked order {signal.symbol} — {elapsed:.0f}min since last exit")
                    return
            log.info(f"[DAY TRADE ENTRY] Opening {signal.symbol} {signal.side.value} position (time: {now_t.strftime('%H:%M:%S')})")
            self.orders.enter(signal)
            self.strategy.record_entry(signal, now)
            
            # Send Telegram entry notification
            self.telegram.send_trade_entry(
                symbol=signal.symbol, side=signal.side.value, strategy=signal.strategy.value,
                entry_price=signal.entry, size=signal.size, account_size=self.strategy.account_size
            )

    def _fetch_prev_closes(self, watchlist):
        log.info("Fetching previous closes for gap detection...")
        for sym in watchlist:
            try:
                df = self.data.fetch_bars(sym, duration="2 D", bar_size="1 day")
                if not df.empty and len(df) >= 2:
                    self.strategy.prev_closes[sym] = float(df.iloc[-2]["close"])
                    log.info(f"  Prev close {sym}: {self.strategy.prev_closes[sym]:.2f}")
            except Exception as e:
                log.warning(f"  Could not fetch prev close for {sym}: {e}")
        log.info("Prev closes fetched. Starting main loop...")

    def _flatten_all(self, reason: ExitReason):
        """STRICT EOD settle — cancel all working orders, flatten every IB position
        (not just bot-tracked), verify, retry, and alert on any survivors."""
        # 1) Cancel ALL open/working orders first so resting stops/limits can't
        #    re-fire after we flatten.
        try:
            open_trades = list(self.ib.openTrades())
            if open_trades:
                log.info(f"[EOD STRICT] Cancelling {len(open_trades)} working orders...")
                for tr in open_trades:
                    try:
                        self.ib.cancelOrder(tr.order)
                    except Exception as e:
                        log.error(f"Failed to cancel order {tr.order.orderId}: {e}")
                self.ib.sleep(2)
        except Exception as e:
            log.error(f"[EOD STRICT] openTrades() failed: {e}")

        # 2) Flatten every position the BOT knows about.
        for sym in list(self.strategy.positions.keys()):
            try:
                pos = self.strategy.positions[sym]
                # Get real-time last price (not stale bar close)
                price = 0.0
                try:
                    from ib_insync import Stock as _Stock
                    _contract = _Stock(sym, "SMART", "USD")
                    self.ib.qualifyContracts(_contract)
                    ticker = self.ib.reqMktData(_contract, "", False, False)
                    self.ib.sleep(2)
                    price = ticker.last or ticker.close or ticker.bid or 0.0
                    self.ib.cancelMktData(_contract)
                except Exception as _pe:
                    log.warning(f"[EOD] reqMktData failed for {sym}: {_pe} — falling back to bars")
                if not price:
                    df = self.data.fetch_bars(sym)
                    price = float(df.iloc[-1]["close"]) if not df.empty else pos.entry_price
                    log.warning(f"[EOD] Using bar close for {sym} price: {price}")
                self.orders.exit(sym, pos.side, pos.size, reason.value)
                self.strategy.record_exit(sym, price, reason)
                pnl = round((price - pos.entry_price) * pos.size * (1 if pos.side == Side.LONG else -1), 2)
                try:
                    self.telegram.send_trade_exit(
                        symbol=sym, side=pos.side.value, strategy=pos.strategy.value,
                        entry_price=pos.entry_price, exit_price=price,
                        size=pos.size, pnl=pnl, reason=reason.value
                    )
                except Exception as e:
                    log.error(f"Failed to send trade exit notification for {sym}: {e}")
            except Exception as e:
                log.error(f"Error flattening {sym}: {e}")

        # 3) Wait for fills, then reconcile against IB's actual positions and
        #    flatten anything still showing — retry up to 3 times.
        from ib_insync import MarketOrder, Stock
        for attempt in range(1, 4):
            self.ib.sleep(5)
            try:
                ib_positions = [p for p in self.ib.positions() if p.position != 0]
                # Only close US stocks (USD currency) — don't touch HK bot's SEHK/HKD positions
                ib_positions = [p for p in ib_positions if p.contract.currency == "USD"]
            except Exception as e:
                log.error(f"[EOD STRICT] positions() failed (attempt {attempt}): {e}")
                continue
            if not ib_positions:
                log.info("[EOD STRICT] ✅ All positions confirmed flat at IB.")
                break
            log.warning(f"[EOD STRICT] Attempt {attempt}: {len(ib_positions)} residual position(s) at IB → force-closing")
            for p in ib_positions:
                try:
                    sym = p.contract.symbol
                    qty = abs(int(p.position))
                    action = "SELL" if p.position > 0 else "BUY"
                    # Use actual IB contract to preserve exchange/currency for non-US stocks
                    contract = p.contract
                    self.ib.qualifyContracts(contract)
                    self.ib.placeOrder(contract, MarketOrder(action, qty))
                    log.info(f"[EOD STRICT] Residual close: {action} {qty} {sym} ({contract.exchange}/{contract.currency})")
                except Exception as e:
                    log.error(f"[EOD STRICT] Failed to close residual {p.contract.symbol}: {e}")
        else:
            # Loop exhausted without break → still residual
            try:
                survivors = [(p.contract.symbol, p.position) for p in self.ib.positions() if p.position != 0]
                if survivors:
                    msg = f"🚨 EOD SETTLE FAILED — survivors: {survivors}"
                    log.error(msg)
                    try:
                        self.telegram.send_message(msg)
                    except Exception:
                        pass
            except Exception:
                pass

        # 4) Final sweep: cancel any orders that re-spawned during flatten.
        try:
            for tr in self.ib.openTrades():
                try:
                    self.ib.cancelOrder(tr.order)
                except Exception:
                    pass
        except Exception:
            pass

    def _print_summary(self):
        log.info(self.strategy.get_summary())
    
    def _send_daily_summary(self):
        """Send end of day summary to Telegram"""
        try:
            wins = sum(s.wins for s in self.strategy.stats.values())
            losses = sum(s.losses for s in self.strategy.stats.values())
            trades = wins + losses or self.strategy.trade_count
            total_pnl = self.strategy.daily_pnl
            
            self.telegram.send_daily_summary(
                trades=trades, wins=wins, losses=losses,
                total_pnl=total_pnl, account_size=self.strategy.account_size
            )
        except Exception as e:
            log.error(f"Error sending daily summary: {e}")

    def _cleanup(self):
        if self.ib.isConnected():
            try:
                self.ib.disconnect()
            except Exception:
                pass
        log.info("Disconnected from IB.")


# ─── Entry Point ──────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="IB Multi-Strategy Algo Trader")
    parser.add_argument("--dry-run", action="store_true", help="Signals only, no real orders")
    parser.add_argument("--port", type=int, default=None)
    args = parser.parse_args()

    if args.port:
        cfg.IB_PORT = args.port

    print(r"""
    ╔═══════════════════════════════════════════════╗
    ║   IB Multi-Strategy Day Trader                ║
    ║                                               ║
    ║   Strategies: ORB | VWAP Rev | Momentum      ║
    ║               Gap Fill | Power Hour           ║
    ║   Risk: Kelly | VIX Regime | Auto-reconnect  ║
    ╚═══════════════════════════════════════════════╝
    """)

    trader = Trader(dry_run=args.dry_run)
    trader.connect(retry=True)
    trader.run()


if __name__ == "__main__":
    main()
